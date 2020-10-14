import os
from typing import List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from back.mergers.excel.cell_util import index_to_cell_name, search_cell_by_column_and_value, cell_above
from back.mergers.excel.student import Student
from back.mergers.merger import Merger
from util.logger import logger

OUTPUT_FILE_NAME = 'Merged.xlsx'
EXCEL_EXTENSION = '.xlsx'

FIRST_STUDENT_NUMBER = 1
FIRST_CELL_SEARCH_LIMIT = 50
LAST_CELL_SEARCH_LIMIT = 150

STUDENT_NUMBER_COLUMN_INDEX = 1
STUDENT_USERNAME_COLUMN_INDEX = 2
STUDENT_FIRST_OVERVIEW_COLUMN = 4

ERROR_MISSING_REVIEW = 'missing_review'
ERROR_VIOLATING_REVIEW = 'violating_review'


# Helper methods
def _get_first_cell_of_sheet(sheet: Worksheet) -> Cell:
    return search_cell_by_column_and_value(sheet, STUDENT_NUMBER_COLUMN_INDEX, FIRST_STUDENT_NUMBER,
                                           FIRST_CELL_SEARCH_LIMIT)


def _get_last_cell_of_sheet(sheet: Worksheet) -> Cell:
    empty_cell = search_cell_by_column_and_value(sheet, STUDENT_NUMBER_COLUMN_INDEX, None, LAST_CELL_SEARCH_LIMIT)
    return cell_above(sheet, empty_cell)


def _is_feedback_list_empty(feedback_list: List) -> bool:
    return set(feedback_list) == {None}


class ExcelMerger(Merger):

    def __init__(self):
        super().__init__()
        self._students: Dict = {}
        self._wb: Workbook
        self._overview_sheet: Worksheet
        self._feedback_sheet: Worksheet
        self._first_overview_cell: Cell
        self._last_overview_cell: Cell
        self._first_feedback_cell: Cell

    def merge(self):
        self._load_document()
        self._load_students()
        self._load_feedback()
        self._merge_feedback()
        self._save_workbook()

    def _load_document(self):
        self._wb = load_workbook(self.file, data_only=True)
        self._overview_sheet = self._wb[self.overview_sheet_name]
        self._feedback_sheet = self._wb[self.feedback_sheet_name]
        self._first_overview_cell = _get_first_cell_of_sheet(self._overview_sheet)
        self._last_overview_cell = _get_last_cell_of_sheet(self._overview_sheet)
        self._first_feedback_cell = _get_first_cell_of_sheet(self._feedback_sheet)
        self._num_students = 1 + self._last_overview_cell.row - self._first_overview_cell.row
        logger.write_log(self._caller, f'Located {self._num_students} students in the overview sheet')

    def _load_students(self):
        for row_index in range(self._first_overview_cell.row, self._num_students + self._first_overview_cell.row):
            num_cell: Cell = self._overview_sheet[index_to_cell_name(STUDENT_NUMBER_COLUMN_INDEX, row_index)]
            username_cell: Cell = self._overview_sheet[index_to_cell_name(STUDENT_USERNAME_COLUMN_INDEX, row_index)]
            student = Student(num_cell.value, username_cell.value)

            for task_index in range(self.num_tasks):
                students_to_review: List = []
                for task_to_review in range(self.num_reviewers_per_task):
                    column_index: int = STUDENT_FIRST_OVERVIEW_COLUMN + task_index * self.num_reviewers_per_task \
                                        + task_to_review
                    student_cell: Cell = self._overview_sheet[index_to_cell_name(column_index, row_index)]
                    students_to_review.append(student_cell.value)
                student.add_students_to_review(task_index, students_to_review)

            self._students[num_cell.value] = student

    def _load_feedback(self):
        for reviewer_num in self._students:
            # Load workbook
            reviewer: Student = self._students[reviewer_num]
            reviewer_feedback_sheet: Worksheet = self._get_student_feedback_sheet(reviewer)
            if not reviewer_feedback_sheet:
                continue

            # Load feedback and insert into student objects
            for task_index in range(self.num_tasks):
                for row in range(self._first_feedback_cell.row, self._num_students + self._first_feedback_cell.row):
                    feedback: List = self._get_feedback(reviewer_feedback_sheet, row, task_index)
                    reviewee_num: int = self.get_student_num(row)
                    reviewee: Student = self._students[reviewee_num]
                    # If student are to be reviewed
                    if reviewee_num in reviewer.students_to_review[task_index]:
                        if _is_feedback_list_empty(feedback):
                            reviewer.add_missing_review(task_index, reviewee_num)
                            self._log_review_error(ERROR_MISSING_REVIEW, task_index, reviewer, reviewee)
                        else:
                            reviewee.add_feedback_from_students(task_index, feedback)
                    # If student shouldn't be reviewed
                    else:
                        if not _is_feedback_list_empty(feedback):
                            reviewer.add_violating_review(task_index, reviewee_num)
                            self._log_review_error(ERROR_VIOLATING_REVIEW, task_index, reviewer, reviewee)

                    pass

    def _merge_feedback(self):
        for student_num in range(FIRST_STUDENT_NUMBER, self._num_students + 1):
            student: Student = self._students[student_num]
            self._write_student_feedback(student)

    def _save_workbook(self):
        merge_file_path = self.output_folder + "\\Merged.xlsx"
        try:
            self._wb.save(merge_file_path)
        except PermissionError as error:
            logger.write_error(self._caller, f'EXCEPTION - Cannot save merged file due to permission error. '
                                             f'Is the file open?')
            raise

    ##############################
    # Class level helper methods #
    ##############################
    def _get_student_workbook(self, username: str) -> Workbook:
        student_file_folder = f'{self.input_folder}\\{username}'
        student_file_path = ''
        for file in os.listdir(student_file_folder):
            if file.endswith(EXCEL_EXTENSION):
                student_file_path = f'{student_file_folder}\\{file}'

        try:
            student_file = load_workbook(student_file_path, data_only=True)
        except PermissionError as error:
            logger.write_error(self._caller, f'EXCEPTION - Cannot open student file ({username}) due to '
                                             f'permission error. Is the file open?')
            raise

        return student_file

    def _get_student_feedback_sheet(self, reviewer: Student) -> Worksheet:
        student_feedback_sheet = None
        try:
            reviewer_file = self._get_student_workbook(reviewer.username)
            student_feedback_sheet = reviewer_file[self.feedback_sheet_name]
        except FileNotFoundError as error:
            logger.write_log(self._caller, f'Student ({reviewer.username}) did not submit a review')

        return student_feedback_sheet

    def _get_feedback(self, feedback_sheet: Worksheet, row: int, task_index: int) -> List:
        feedback_row: int = row
        feedback_start_column: int = STUDENT_FIRST_OVERVIEW_COLUMN + task_index * self.columns_per_task

        feedback: List = []
        for cell_column in range(self.columns_per_task):
            feedback_column: int = feedback_start_column + cell_column
            feedback_cell: Cell = feedback_sheet[index_to_cell_name(feedback_column, feedback_row)]
            feedback.append(feedback_cell.value)

        return feedback

    def get_student_num(self, row):
        return row - self._first_feedback_cell.row + 1

    def _write_student_feedback(self, student: Student):
        """
            Write the feedback of a student into the feedback sheet of the output file

        :param student: the student object which holds the feedback
        :return:
        """

        row: int = self._get_row(student.number)
        for feedback_item in student.feedback:
            task_index: int = feedback_item['task_index']
            task_feedback: List = feedback_item['feedback']
            self._write_task_feedback(row, task_index, task_feedback)

    def _write_task_feedback(self, row: int, index: int, feedback: List):
        """
            Write the feedback of a task starting from the start_column

        :param row: row in sheet of output file
        :param index: task index
        :param feedback: feedback to be written
        :return:
        """
        start_column: int = STUDENT_FIRST_OVERVIEW_COLUMN + index * self.num_reviewers_per_task * self.columns_per_task

        # Move to next set of columns if there exists multiple sets of feedback for the student's task
        while self._is_task_reviewed(start_column, row) is True:
            start_column = start_column + self.columns_per_task

        for feedback_index in range(self.columns_per_task):
            column: int = start_column + feedback_index

            feedback_cell: Cell = self._feedback_sheet[index_to_cell_name(column, row)]
            feedback_cell.value = feedback[feedback_index]

    def _is_task_reviewed(self, column, row):
        for feedback_index in range(self.columns_per_task):
            feedback_column: int = column + feedback_index
            feedback_cell: Cell = self._feedback_sheet[index_to_cell_name(feedback_column, row)]
            if feedback_cell.value is not None:
                return True

        return False

    def _get_row(self, student_num: int) -> int:
        return self._first_feedback_cell.row + student_num - 1

    def _log_review_error(self, error: str, task_index: int, reviewer: Student, reviewee: Student):
        if error is ERROR_MISSING_REVIEW:
            logger.write_error(self._caller, f'{reviewer.username} missing review in task {task_index} '
                                             f'for {reviewee.username} ')
        elif error is ERROR_VIOLATING_REVIEW:
            logger.write_error(self._caller, f'{reviewer.username} has input in task {task_index} '
                                             f'for {reviewee.username} violating reviewing schema')
